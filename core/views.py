import json
from datetime import date, datetime, timedelta
from decimal import Decimal

import openpyxl
from escpos.printer import Usb as EscposUsb

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.db import transaction
from django.db.models import (
    Count,
    DecimalField,
    ExpressionWrapper,
    F,
    FloatField,
    Q,
    Sum,
    Value,
)
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date

from .decorators import es_admin, es_vendedor
from .forms import AbonoForm, ClienteForm, ProductForm
from .models import Abono, Cliente, Product, Venta, VentaItem
from django.contrib.auth.models import Group


def es_vendedor(user):
    return user.is_authenticated and user.groups.filter(name="vendedor").exists()


def es_admin(user):
    return user.is_authenticated and user.groups.filter(name="admin").exists()


# --- Login y redirección por rol ---
def login_view(request):
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.get_user()
        auth_login(request, user)
        return redirect("redirect_by_role")
    return render(request, "core/login.html", {"form": form})


@login_required
def redirect_by_role(request):
    if request.user.is_superuser:
        return redirect("admin_dashboard")
    return redirect("vendedor_dashboard")


# core/views.py
@login_required
@user_passes_test(es_admin, login_url="login")
def admin_dashboard(request):
    """
    Panel de Control para administradores:
    • Clientes por vendedor
    • Ingresos del mes
    • Productos con stock bajo
    • Gráficos de ventas semanales y proporción de tipos de pago
    """
    hoy = date.today()

    # KPI: Clientes por vendedor
    clientes_por_vendedor = Cliente.objects.values("creado_por__username").annotate(
        total=Count("id")
    )

    # Gráfico 1: Ventas últimos 7 días
    fechas = [hoy - timedelta(days=i) for i in range(6, -1, -1)]
    ventas_semana_qs = (
        Venta.objects.filter(fecha__date__gte=hoy - timedelta(days=6), estado="activa")
        .annotate(dia=TruncDate("fecha"))
        .values("dia")
        .annotate(total_dia=Sum("total"))
        .order_by("dia")
    )
    mapa_semana = {v["dia"].strftime("%d/%m"): v["total_dia"] for v in ventas_semana_qs}
    fechas_semana = [d.strftime("%d/%m") for d in fechas]
    # Convertimos a float para que json.dumps funcione
    serie_ventas = [float(mapa_semana.get(label, 0)) for label in fechas_semana]

    # Gráfico 2: Proporción de tipos de pago (últ. 30 días)
    pagos_qs = (
        Venta.objects.filter(fecha__date__gte=hoy - timedelta(days=30))
        .values("tipo_pago")
        .annotate(cantidad=Count("id"))
    )
    labels_pagos = [p["tipo_pago"].capitalize() for p in pagos_qs]
    datos_pagos = [p["cantidad"] for p in pagos_qs]  # son ints, no hay problema

    # KPI: Ingresos del mes actual
    ingresos_mes = (
        Venta.objects.filter(
            fecha__year=hoy.year, fecha__month=hoy.month, estado="activa"
        ).aggregate(total=Sum("total"))["total"]
        or 0
    )
    # ingresos_mes también es Decimal, pero lo mostramos en plantilla, no lo serializamos a JSON

    # KPI: Productos con stock bajo
    bajo_stock = Product.objects.filter(stock__lte=5).count()

    # Serializamos las listas a JSON ya limpias
    context = {
        "clientes_por_vendedor": clientes_por_vendedor,
        "ingresos_mes": ingresos_mes,
        "bajo_stock": bajo_stock,
        "fechas_semana_json": json.dumps(fechas_semana),
        "serie_ventas_json": json.dumps(serie_ventas),
        "labels_pagos_json": json.dumps(labels_pagos),
        "datos_pagos_json": json.dumps(datos_pagos),
    }

    return render(request, "core/admin/admin_dashboard.html", context)


# core/views.py
@login_required
@user_passes_test(es_vendedor, login_url="login")
def vendedor_dashboard(request):
    hoy = date.today()

    # — Ventas de hoy —
    ventas_hoy_qs = Venta.objects.filter(
        usuario=request.user, fecha__date=hoy, estado="activa"
    )
    ventas_hoy_count = ventas_hoy_qs.count()
    ventas_hoy_total = ventas_hoy_qs.aggregate(total=Sum("total"))["total"] or 0

    # — Abonos pendientes (ventas a crédito con saldo > 0) —
    ventas_credito = Venta.objects.filter(
        usuario=request.user, tipo_pago="credito", estado="activa"
    )
    abonos_pendientes = []
    for v in ventas_credito:
        saldo = v.calcular_saldo_pendiente()
        if saldo > 0:
            abonos_pendientes.append(
                {
                    "factura": v.numero_factura,
                    "cliente": v.cliente.nombre,
                    "saldo": saldo,
                }
            )
    abonos_pendientes_count = len(abonos_pendientes)

    # — Productos agotándose (stock ≤5) de los que ha vendido —
    prod_ids_vend = (
        VentaItem.objects.filter(venta__usuario=request.user)
        .values_list("producto_id", flat=True)
        .distinct()
    )
    productos_criticos = Product.objects.filter(
        id__in=prod_ids_vend, stock__lte=5
    ).values("reference", "description", "stock")
    productos_agotandose_count = productos_criticos.count()

    # — Top 5 clientes por ventas del mes —
    primer_dia_mes = hoy.replace(day=1)
    clientes_mes = (
        Venta.objects.filter(
            usuario=request.user, fecha__date__gte=primer_dia_mes, estado="activa"
        )
        .values("cliente__nombre")
        .annotate(total_mes=Sum("total"))
        .order_by("-total_mes")[:5]
    )
    top_clients = [
        {"nombre": c["cliente__nombre"], "total": c["total_mes"]} for c in clientes_mes
    ]

    # — Productos más vendidos última semana —
    semana_atras = hoy - timedelta(days=6)
    productos_semana_qs = (
        VentaItem.objects.filter(
            venta__usuario=request.user,
            venta__fecha__date__gte=semana_atras,
            venta__estado="activa",
        )
        .values("producto__reference", "producto__description")
        .annotate(cantidad=Sum("cantidad"))
        .order_by("-cantidad")[:5]
    )
    productos_mas_vendidos = [
        {
            "reference": p["producto__reference"],
            "description": p["producto__description"],
            "cantidad": p["cantidad"],
        }
        for p in productos_semana_qs
    ]

    # — Gráfico Ventas últimos 7 días —
    fechas = [hoy - timedelta(days=i) for i in range(6, -1, -1)]
    ventas_semana = (
        Venta.objects.filter(
            usuario=request.user, fecha__date__gte=semana_atras, estado="activa"
        )
        .annotate(dia=TruncDate("fecha"))
        .values("dia")
        .annotate(total_dia=Sum("total"))
        .order_by("dia")
    )
    mapa = {v["dia"].strftime("%d/%m"): float(v["total_dia"]) for v in ventas_semana}
    fechas_semana = [d.strftime("%d/%m") for d in fechas]
    serie_ventas = [mapa.get(label, 0) for label in fechas_semana]

    # — Gráfico Tipos de pago (últ. 30 días) —
    pagos_qs = (
        Venta.objects.filter(
            usuario=request.user, fecha__date__gte=hoy - timedelta(days=30)
        )
        .values("tipo_pago")
        .annotate(cantidad=Count("id"))
    )
    labels_pagos = [p["tipo_pago"].capitalize() for p in pagos_qs]
    datos_pagos = [p["cantidad"] for p in pagos_qs]

    # Serializamos los arrays para Chart.js
    context = {
        "ventas_hoy_count": ventas_hoy_count,
        "ventas_hoy_total": ventas_hoy_total,
        "abonos_pendientes": abonos_pendientes,
        "abonos_pendientes_count": abonos_pendientes_count,
        "productos_criticos": productos_criticos,
        "productos_agotandose_count": productos_agotandose_count,
        "top_clients": top_clients,
        "productos_mas_vendidos": productos_mas_vendidos,
        "fechas_semana_json": json.dumps(fechas_semana),
        "serie_ventas_json": json.dumps(serie_ventas),
        "labels_pagos_json": json.dumps(labels_pagos),
        "datos_pagos_json": json.dumps(datos_pagos),
    }
    return render(request, "core/vendedor/vendedor_dashboard.html", context)


# --- CRUD Productos ---
@login_required
@user_passes_test(es_admin, login_url="login")
def admin_product_list(request):
    query = request.GET.get("buscar", "").strip()
    qs = Product.objects.all()
    if query:
        qs = qs.filter(Q(reference__icontains=query) | Q(description__icontains=query))
    products = qs.order_by("reference")

    # Total de stock
    total_stock = products.aggregate(total_stock=Sum("stock"))["total_stock"] or 0

    # Total valor de inventario = sum(purchase_price * stock)
    total_valor_inventario = (
        products.aggregate(
            total_valor=Sum(
                F("purchase_price") * F("stock"),
                output_field=FloatField(),  # <- aquí lo usas
            )
        )["total_valor"]
        or 0
    )

    return render(
        request,
        "core/admin/products/list.html",
        {
            "products": products,
            "query": query,
            "total_stock": total_stock,
            "total_valor_inventario": total_valor_inventario,
        },
    )


def vendedor_product_list(request):
    query = request.GET.get("buscar", "").strip()
    products = (
        Product.objects.filter(
            Q(reference__icontains=query) | Q(description__icontains=query)
        )
        if query
        else Product.objects.all()
    )

    return render(
        request,
        "core/vendedor/products/list.html",  # plantilla específica para vendedor
        {
            "products": products,
        },
    )


def product_create(request):
    form = ProductForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("admin_product_list")  # Asegúrate de que esta URL exista
    return render(request, "core/admin/products/create.html", {"form": form})


def product_edit(request, producto_id):
    producto = get_object_or_404(Product, id=producto_id)

    if request.method == "POST":
        reference = request.POST.get("reference", "").strip()
        description = request.POST.get("description", "").strip()
        purchase_price = request.POST.get("purchase_price")
        stock_add = request.POST.get("stock_add")

        if not reference or not description:
            messages.error(request, "Todos los campos son obligatorios.")
        else:
            try:
                producto.reference = reference
                producto.description = description
                producto.purchase_price = float(purchase_price)
                producto.stock += int(stock_add)
                producto.save()

                messages.success(request, "✅ Producto actualizado correctamente.")
                return redirect(
                    "admin_product_list"
                )  # Cambia esto si el nombre de la URL es diferente
            except Exception as e:
                messages.error(request, f"Error al actualizar el producto: {str(e)}")

    return render(request, "core/admin/products/edit.html", {"producto": producto})


def product_delete(request, pk):
    prod = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        prod.delete()
        return redirect("admin_product_list")
    return render(request, "core/admin/products/delete.html", {"product": prod})


@login_required
@user_passes_test(es_vendedor, login_url="login")
def vendedor_cliente_list(request):
    query = request.GET.get("q", "").strip()
    clientes = (
        Cliente.objects.filter(Q(nombre__icontains=query) | Q(cedula__icontains=query))
        if query
        else Cliente.objects.all()
    )

    return render(
        request,
        "core/vendedor/clientes/list.html",
        {"clientes": clientes, "query": query},
    )


# core/views.py
@login_required
@user_passes_test(es_admin, login_url="login")
def admin_cliente_list(request):
    query = request.GET.get("q", "").strip()
    if query:
        clientes = Cliente.objects.filter(
            Q(nombre__icontains=query)
            | Q(cedula__icontains=query)
            | Q(email__icontains=query)
        )
    else:
        clientes = Cliente.objects.all()
    return render(
        request,
        "core/admin/clientes/list.html",
        {
            "clientes": clientes,
            "query": query,
        },
    )


# core/views.py
@login_required
def cliente_create(request):
    # Inicializamos el formulario con POST o en blanco
    form = ClienteForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.creado_por = request.user
            cliente.save()
            # Redirige según rol
            if request.user.is_superuser:
                return redirect("admin_cliente_list")
            return redirect("vendedor_cliente_list")
        # Si POST pero inválido, caerá al render de abajo mostrando errores

    # Para GET o POST inválido, renderizamos el formulario
    template = (
        "core/admin/clientes/create.html"
        if request.user.is_superuser
        else "core/vendedor/clientes/create.html"
    )
    return render(request, template, {"form": form})


def cliente_historial(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    ventas = Venta.objects.filter(cliente=cliente)
    return render(
        request,
        "core/admin/clientes/historial.html",
        {"cliente": cliente, "ventas": ventas},
    )


@login_required
def cliente_edit(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)

    if request.method == "POST":
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            if request.user.is_superuser:
                return redirect("admin_cliente_list")
            return redirect("vendedor_cliente_list")
    else:
        form = ClienteForm(instance=cliente)

    return render(
        request, "core/admin/clientes/edit.html", {"form": form, "cliente": cliente}
    )


@login_required
def cliente_delete(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if request.method == "POST":
        cliente.delete()
        if request.user.is_superuser:
            return redirect("admin_cliente_list")
        return redirect("vendedor_cliente_list")
    return render(request, "core/admin/clientes/delete.html", {"cliente": cliente})


# --- Ventas ---
@login_required
def buscar_clientes(request):
    term = request.GET.get("q") or request.GET.get("term") or ""
    clientes = Cliente.objects.filter(nombre__icontains=term)
    return JsonResponse(
        [{"id": c.id, "nombre": c.nombre, "cedula": c.cedula} for c in clientes],
        safe=False,
    )


def buscar_productos(request):
    term = request.GET.get("q") or request.GET.get("term") or ""
    productos = Product.objects.filter(
        Q(reference__icontains=term) | Q(description__icontains=term)
    )
    return JsonResponse(
        [
            {
                "id": p.id,
                "reference": p.reference,
                "description": p.description,
                "stock": p.stock,  # Agregado aquí
            }
            for p in productos
        ],
        safe=False,
    )


# ventas
# core/views.py
def _abrir_impresora_usb():
    """
    Intenta detectar cualquier impresora USB de clase impresora (bDeviceClass == 7).
    Si no hay PyUSB o no encuentra ninguna, cae al fallback definido en settings.
    """
    try:
        import usb.core
    except ImportError:
        usb = None

    if usb:
        for dev in usb.core.find(
            find_all=True, custom_match=lambda d: d.bDeviceClass == 7
        ):
            try:
                return EscposUsb(
                    dev.idVendor, dev.idProduct, timeout=settings.ESC_POS_USB_TIMEOUT
                )
            except Exception:
                continue

    # Fallback a los IDs en settings
    return EscposUsb(
        settings.ESC_POS_USB_VENDOR,
        settings.ESC_POS_USB_PRODUCT,
        timeout=settings.ESC_POS_USB_TIMEOUT,
    )


@login_required
@user_passes_test(es_vendedor, login_url="login")
def crear_venta(request):
    if request.method == "GET":
        return render(
            request,
            "core/vendedor/ventas/create.html",
            {
                "clientes": Cliente.objects.all(),
                "productos": Product.objects.values("id", "reference", "description"),
            },
        )

    data = request.POST
    cliente_id = data.get("cliente")
    productos_json = json.loads(data.get("productos_data", "[]"))
    descuento = Decimal(data.get("descuento", "0").replace(".", ""))
    tipo_pago = data.get("tipo_pago", "contado")

    if not cliente_id or not productos_json:
        messages.error(request, "Debes seleccionar un cliente y al menos un producto.")
        return redirect("venta_create")

    try:
        with transaction.atomic():
            pref = {"credito": "FC-", "transferencia": "FT-", "garantia": "FG-"}.get(
                tipo_pago, "FV-"
            )
            nro = Venta.objects.filter(tipo_pago=tipo_pago).count() + 1
            venta = Venta.objects.create(
                cliente_id=cliente_id,
                numero_factura=f"{pref}{nro}",
                tipo_pago=tipo_pago,
                usuario=request.user,
            )

            bruto = Decimal(0)
            for itm in productos_json:
                prod = Product.objects.get(pk=itm["producto_id"])
                cant = int(itm["cantidad"])
                precio = Decimal(str(itm.get("precio", "0")))

                if prod.stock < cant:
                    raise ValueError(f"Stock insuficiente {prod.reference}")

                sub = precio * cant
                VentaItem.objects.create(
                    venta=venta, producto=prod, cantidad=cant, precio=precio
                )
                prod.stock -= cant
                prod.save()
                bruto += sub

            venta.total = max(bruto - descuento, Decimal(0))
            venta.save()

        messages.success(request, f"✅ Venta #{venta.numero_factura} registrada.")
        return redirect("venta_vendedor_list")

    except Exception as e:
        messages.error(request, f"❌ Error al registrar la venta: {e}")
        return redirect("venta_create")


@login_required
@user_passes_test(es_vendedor, login_url="login")
def imprimir_venta(request, venta_id):
    """
    Vista legacy de impresión directa via USB.
    La dejamos por compatibilidad, pero recomendamos usar `ticket_venta`.
    """
    venta = get_object_or_404(Venta, pk=venta_id)
    try:
        p = _abrir_impresora_usb()
        p.set(align="center")
        p.text("ZONA T\n")
        p.set(align="left")
        p.text(
            f"{venta.cliente.nombre}\n{venta.cliente.direccion}\n{venta.cliente.telefono}\n"
        )
        p.text(
            f"Venta N° {venta.numero_factura}\nMétodo: {venta.tipo_pago.capitalize()}\n"
        )
        p.text(f"Fecha: {timezone.localtime(venta.created_at):%d/%m/%Y %H:%M}\n")
        p.text("-" * 32 + "\n")
        p.text("REF   DESC       CANT  VALOR\n")
        p.text("-" * 32 + "\n")
        for item in venta.items.all():
            ref = item.producto.reference[:10]
            desc = item.producto.description[:10]
            p.text(f"{ref:10s} {desc:10s} {item.cantidad:3d} {item.precio:7.2f}\n")
        p.text("-" * 32 + "\n")
        p.text(f"TOTAL: {venta.total:.2f}\n")
        p.cut()
        messages.success(request, "🖨️ Tirilla enviada a la impresora.")
    except Exception as e:
        messages.warning(request, f"❗ Venta registrada, pero NO se imprimió: {e}")
    return redirect("venta_vendedor_list")


@login_required
@user_passes_test(
    lambda u: u.is_authenticated and (es_vendedor(u) or es_admin(u)), login_url="login"
)
def ticket_venta(request, venta_id):
    """
    Genera un ticket HTML con @page size:80mm y window.print(),
    reutilizable por vendedor y administrador.
    """
    venta = get_object_or_404(Venta, pk=venta_id)
    return render(request, "core/print.html", {"venta": venta})


@login_required
@user_passes_test(es_admin, login_url="login")
def venta_admin_detail(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id)
    return render(request, "core/admin/ventas/detail.html", {"venta": venta})


@login_required
@user_passes_test(es_vendedor, login_url="login")
def venta_vendedor_detail(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id, usuario=request.user)
    return render(request, "core/vendedor/ventas/detail.html", {"venta": venta})


# core/views.py
@login_required
@user_passes_test(es_admin, login_url="login")
def venta_admin_list(request):
    query = request.GET.get("q", "").strip()
    tipo_pago = request.GET.get("tipo_pago", "")
    fecha_inicio = parse_date(
        request.GET.get("fecha_inicio", "")
    )  # datetime.date o None
    fecha_fin = parse_date(request.GET.get("fecha_fin", ""))

    # para usar en el template si no seleccionan fecha_inicio:
    hoy = date.today()
    export_fecha = (fecha_inicio or hoy).strftime("%Y-%m-%d")

    qs = (
        Venta.objects.select_related("cliente", "usuario")
        .prefetch_related("items__producto", "abonos__usuario")
        .order_by("-fecha")
    )
    if query:
        qs = qs.filter(
            Q(cliente__nombre__icontains=query) | Q(numero_factura__icontains=query)
        )
    if fecha_inicio:
        qs = qs.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__date__lte=fecha_fin)

    ventas = []
    for v in qs:
        v.ganancia = sum(
            (item.precio - item.producto.purchase_price) * item.cantidad
            for item in v.items.all()
        )
        ultimo = v.abonos.order_by("-fecha").first()
        v.ultimo_abono_monto = ultimo.monto if ultimo else None
        v.ultimo_abono_por = (
            ultimo.usuario.username if ultimo and ultimo.usuario else None
        )
        ventas.append(v)

    total_ventas = sum(v.total for v in ventas)
    total_ganancias = sum(v.ganancia for v in ventas)

    return render(
        request,
        "core/admin/ventas/list.html",
        {
            "ventas": ventas,
            "query": query,
            "fecha_inicio": fecha_inicio and fecha_inicio.strftime("%Y-%m-%d"),
            "fecha_fin": fecha_fin and fecha_fin.strftime("%Y-%m-%d"),
            "total_ventas": total_ventas,
            "total_ganancias": total_ganancias,
            "export_fecha": export_fecha,  # ← pasamos este valor al template
        },
    )


# core/views.py
@login_required
@user_passes_test(es_vendedor, login_url="login")
def venta_vendedor_list(request):
    query = request.GET.get("q", "").strip()
    fecha_inicio = request.GET.get("fecha_inicio", "").strip()
    fecha_fin = request.GET.get("fecha_fin", "").strip()

    qs = (
        Venta.objects.filter(usuario=request.user)
        .select_related("cliente")
        .prefetch_related("abonos__usuario")
        .order_by("-fecha")
    )
    if query:
        qs = qs.filter(
            Q(cliente__nombre__icontains=query) | Q(numero_factura__icontains=query)
        )
    if fecha_inicio:
        qs = qs.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__date__lte=fecha_fin)

    ventas = []
    for v in qs:
        # último abono
        ultimo = v.abonos.order_by("-fecha").first()
        v.ultimo_abono_por = (
            ultimo.usuario.username if (ultimo and ultimo.usuario) else None
        )
        v.ultimo_abono_monto = ultimo.monto if ultimo else None
        ventas.append(v)

    total_ventas = sum(v.total for v in ventas)

    return render(
        request,
        "core/vendedor/ventas/list.html",
        {
            "ventas": ventas,
            "query": query,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "total_ventas": total_ventas,
        },
    )


def detalle_venta(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id)
    items = VentaItem.objects.filter(venta=venta)
    return render(
        request,
        "core/admin/ventas/detail.html",
        {
            "venta": venta,
            "items": items,
        },
    )


@login_required
def venta_delete(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id)

    if venta.estado == "anulada":
        messages.warning(request, "Esta venta ya está anulada.")
        return redirect(
            "venta_admin_list" if request.user.is_superuser else "venta_vendedor_list"
        )

    try:
        with transaction.atomic():
            for item in venta.items.all():
                producto = item.producto
                producto.stock += item.cantidad
                producto.save()

            venta.estado = "anulada"
            venta.save()

            messages.success(
                request, "✅ Venta anulada correctamente. El stock fue restaurado."
            )
    except Exception as e:
        messages.error(request, f"❌ Error al anular la venta: {str(e)}")

    return redirect(
        "venta_admin_list" if request.user.is_superuser else "venta_vendedor_list"
    )


# --- Pagos ---
@login_required
def registrar_abono(request):
    venta_id = request.GET.get("venta_id")

    if request.method == "POST":
        form = AbonoForm(request.POST)
        if form.is_valid():
            abono = form.save(commit=False)
            abono.usuario = request.user
            abono.save()
            # Redirige al listado de saldos pendientes según rol
            if request.user.is_superuser:
                return redirect("saldo_pendiente_admin")
            else:
                return redirect("saldo_pendiente")
    else:
        form = AbonoForm(initial={"venta": venta_id}) if venta_id else AbonoForm()

    return render(
        request,
        "core/admin/pagos/registrar_abono.html",  # o core/vendedor/... si tienes plantilla distinta
        {"form": form},
    )


# core/views.py
@login_required
@user_passes_test(es_admin, login_url="login")
def saldo_pendiente_admin(request):
    query = request.GET.get("q", "").strip()

    qs = (
        Venta.objects.filter(tipo_pago="credito")
        .select_related("cliente", "usuario")
        .order_by("-fecha")
    )
    if query:
        qs = qs.filter(
            Q(cliente__nombre__icontains=query) | Q(numero_factura__icontains=query)
        )

    ventas = []
    for v in qs:
        saldo = v.calcular_saldo_pendiente()
        if saldo > 0:
            # datos que ya tenías
            v.total_compra = v.total
            v.saldo_pendiente = saldo

            # extraigo el último abono
            ultimo = v.abonos.order_by("-fecha").first()
            if ultimo:
                v.ultimo_abono_monto = ultimo.monto
                v.ultimo_abono_por = ultimo.usuario.username if ultimo.usuario else None
                v.ultimo_abono_fecha = ultimo.fecha
            else:
                v.ultimo_abono_monto = None
                v.ultimo_abono_por = None
                v.ultimo_abono_fecha = None

            ventas.append(v)

    total_deuda = sum(v.saldo_pendiente for v in ventas)

    return render(
        request,
        "core/admin/pagos/saldo_pendiente.html",
        {
            "ventas": ventas,
            "query": query,
            "total_deuda": total_deuda,
        },
    )


@login_required
@user_passes_test(es_vendedor, login_url="login")
def saldo_pendiente(request):
    query = request.GET.get("q", "").strip()

    qs = (
        Venta.objects.filter(usuario=request.user, tipo_pago="credito")
        .select_related("cliente")
        .order_by("-fecha")
    )
    if query:
        qs = qs.filter(
            Q(cliente__nombre__icontains=query) | Q(numero_factura__icontains=query)
        )

    ventas = []
    for v in qs:
        saldo = v.calcular_saldo_pendiente()
        if saldo > 0:
            v.total_compra = v.total
            v.saldo_pendiente = saldo

            ultimo = v.abonos.order_by("-fecha").first()
            if ultimo:
                v.ultimo_abono_monto = ultimo.monto
                v.ultimo_abono_por = ultimo.usuario.username if ultimo.usuario else None
                v.ultimo_abono_fecha = ultimo.fecha
            else:
                v.ultimo_abono_monto = None
                v.ultimo_abono_por = None
                v.ultimo_abono_fecha = None

            ventas.append(v)

    total_deuda = sum(v.saldo_pendiente for v in ventas)

    return render(
        request,
        "core/vendedor/pagos/saldo_pendiente.html",
        {
            "ventas": ventas,
            "query": query,
            "total_deuda": total_deuda,
        },
    )


# --- Reportes ---
# Reporte 1: Reporte de ventas diarias
@login_required
def reporte_ventas_diarias(request):
    fecha_str = request.GET.get("fecha")
    if fecha_str:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    else:
        fecha = timezone.now().date()

    ventas = (
        Venta.objects.filter(fecha__date=fecha, estado="activa")
        .select_related("cliente", "usuario")
        .prefetch_related("items__producto")
    )

    total_dia = sum(venta.total for venta in ventas)

    return render(
        request,
        "core/admin/reportes/reporte_ventas_diarias.html",
        {
            "ventas": ventas,
            "fecha": fecha.strftime("%Y-%m-%d"),
            "total_dia": total_dia,
        },
    )


# Reporte 2: Clientes con deuda
def reporte_clientes_con_deuda(request):
    query = request.GET.get("q", "")
    clientes = Cliente.objects.all()

    if query:
        clientes = clientes.filter(nombre__icontains=query)

    clientes_con_deuda = []

    for cliente in clientes:
        ventas = cliente.venta_set.filter(estado="activa", tipo_pago="credito")
        saldo_total = sum(venta.calcular_saldo_pendiente() for venta in ventas)
        if saldo_total > 0:
            cliente.deuda_total = saldo_total  # asignar atributo dinámico
            clientes_con_deuda.append(cliente)

    return render(
        request,
        "core/admin/reportes/clientes_con_deuda.html",
        {"clientes": clientes_con_deuda},
    )


# Reporte 3: Productos más vendidos
def reporte_productos_mas_vendidos(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    ventas = VentaItem.objects.filter(venta__estado="activa")

    if fecha_inicio:
        ventas = ventas.filter(venta__fecha__gte=fecha_inicio)
    if fecha_fin:
        ventas = ventas.filter(venta__fecha__lte=fecha_fin)

    productos = (
        ventas.values("producto__description")
        .annotate(total_vendido=Sum("cantidad"))
        .order_by("-total_vendido")
    )

    for p in productos:
        p["nombre"] = p.pop("producto__description")

    return render(
        request,
        "core/admin/reportes/productos_mas_vendidos.html",
        {"productos": productos},
    )


# Exportar
@login_required
@user_passes_test(es_admin, login_url="login")
def exportar_ventas_excel(request, fecha):
    """
    Genera un XLSX con las ventas del día indicado (YYYY-MM-DD en la URL).
    Incluye columna de Ganancia.
    """
    # 1) Parsear la fecha de la URL
    try:
        dia = datetime.strptime(fecha, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Formato de fecha inválido", status=400)

    # 2) Queryset de ventas activas de ese día
    ventas = Venta.objects.filter(fecha__date=dia, estado="activa").prefetch_related(
        "items__producto"
    )

    # 3) Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ventas {fecha}"

    # 4) Cabecera
    headers = [
        "Factura #",
        "Cliente",
        "Tipo Pago",
        "Total Venta",
        "Fecha / Hora",
        "Ganancia",
    ]
    ws.append(headers)

    # 5) Filas
    for v in ventas:
        # calcular ganancia sumando (precio_venta - purchase_price) * cantidad
        ganancia = sum(
            (item.precio - item.producto.purchase_price) * item.cantidad
            for item in v.items.all()
        )
        ws.append(
            [
                v.numero_factura,
                v.cliente.nombre,
                v.tipo_pago.capitalize(),
                float(v.total),  # para que OpenPyXL no reclame Decimal
                v.fecha.strftime("%Y-%m-%d %H:%M"),
                float(ganancia),
            ]
        )

    # 6) Construir respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="ventas_{fecha}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin, login_url="login")
def exportar_ventas_excel_admin(request):
    ventas = Venta.objects.select_related("cliente", "usuario").order_by("-fecha")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas Administrador"

    ws.append(
        ["Factura", "Cliente", "Tipo de Pago", "Total", "Fecha", "Estado", "Usuario"]
    )

    for v in ventas:
        ws.append(
            [
                v.numero_factura,
                v.cliente.nombre,
                v.tipo_pago.capitalize(),
                float(v.total),
                v.fecha.strftime("%Y-%m-%d %H:%M"),
                v.estado.capitalize(),
                v.usuario.username if v.usuario else "—",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ventas_admin.xlsx"'
    wb.save(response)
    return response


def exportar_clientes_deuda_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes con Deuda"

    # Cabeceras
    ws.append(["Nombre", "Teléfono", "Ciudad", "Deuda ($)"])

    clientes = Cliente.objects.all()

    for cliente in clientes:
        ventas = cliente.venta_set.filter(estado="activa", tipo_pago="credito")
        saldo_total = sum(venta.calcular_saldo_pendiente() for venta in ventas)
        if saldo_total > 0:
            ws.append(
                [cliente.nombre, cliente.telefono, cliente.ciudad, float(saldo_total)]
            )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=clientes_con_deuda.xlsx"
    wb.save(response)
    return response


# core/views.py
def exportar_productos_mas_vendidos_excel(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    # 1. Filtra los items de venta según fechas (si las proporcionas)
    items = VentaItem.objects.select_related("producto", "venta")
    if fecha_inicio:
        items = items.filter(venta__fecha__date__gte=fecha_inicio)
    if fecha_fin:
        items = items.filter(venta__fecha__date__lte=fecha_fin)

    # 2. Agrupa por producto__description (campo correcto) y suma cantidades
    productos = (
        items.values("producto__description")
        .annotate(total_vendido=Sum("cantidad"))
        .order_by("-total_vendido")
    )

    # 3. Crea el Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos Más Vendidos"

    # Cabecera
    ws.append(["Producto", "Unidades Vendidas"])
    for p in productos:
        ws.append([p["producto__description"], p["total_vendido"]])

    # 4. Envía el archivo al cliente
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=productos_mas_vendidos.xlsx"
    wb.save(response)
    return response


def reportes_index(request):
    return render(request, "core/admin/reportes/index.html")


def vista_reportes(request):
    return render(request, "core/admin/reportes/reportes_inicio.html")


# carga masiva


def limpiar_valor(valor):
    if isinstance(valor, (int, float)):
        return valor
    valor_str = (
        str(valor)
        .replace("$", "")
        .replace(",", "")
        .replace("(", "-")
        .replace(")", "")
        .strip()
    )
    try:
        return float(valor_str)
    except ValueError:
        return 0  # O puedes usar: raise ValueError("Valor inválido en Excel")


@login_required
@user_passes_test(es_vendedor, login_url="login")
def exportar_ventas_excel_vendedor(request):
    ventas = (
        Venta.objects.filter(usuario=request.user)
        .select_related("cliente")
        .order_by("-fecha")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mis Ventas"

    # Encabezados
    ws.append(["Factura", "Cliente", "Tipo de Pago", "Total", "Fecha", "Estado"])

    for v in ventas:
        ws.append(
            [
                v.numero_factura,
                v.cliente.nombre,
                v.tipo_pago.capitalize(),
                float(v.total),
                v.fecha.strftime("%Y-%m-%d %H:%M"),
                v.estado.capitalize(),
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="mis_ventas.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin, login_url="login")
def cargar_productos_excel(request):
    """
    Sube un XLSX con columnas:
      1) referencia
      2) descripción
      3) valor_compra (puede venir con comas/miles)
      4) stock
    """
    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        try:
            wb = openpyxl.load_workbook(archivo)
        except Exception:
            messages.error(request, "No pude leer el archivo Excel.")
            return redirect("admin_product_list")

        hoja = wb.active
        filas_procesadas = 0
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            # Si la fila está completamente vacía, la saltamos
            if not any(fila):
                continue

            # Asegurarnos de que tenga al menos 4 celdas
            if len(fila) < 4:
                messages.warning(
                    request,
                    f"Saltada fila {hoja.iter_rows().index(fila)+1}: columnas insuficientes.",
                )
                continue

            referencia, descripcion, valor_compra_raw, stock_raw = fila

            # Limpiar y parsear los valores numéricos
            try:
                valor_compra = limpiar_valor(valor_compra_raw)
                stock = int(limpiar_valor(stock_raw))
            except Exception:
                messages.warning(
                    request,
                    f"Saltada fila con referencia {referencia}: datos numéricos inválidos.",
                )
                continue

            if stock < 0:
                messages.warning(request, f"Saltada fila {referencia}: stock negativo.")
                continue

            # Actualiza o crea el producto (sin tocar sale_price)
            Product.objects.update_or_create(
                reference=referencia,
                defaults={
                    "description": descripcion,
                    "purchase_price": valor_compra,
                    "stock": stock,
                },
            )
            filas_procesadas += 1

        messages.success(
            request,
            f"{filas_procesadas} productos cargados/actualizados correctamente.",
        )
        return redirect("admin_product_list")

    return render(request, "core/admin/products/cargar_productos.html")


# core/views.py
@login_required
@user_passes_test(es_admin, login_url="login")
def exportar_clientes_excel(request):
    """
    Exporta a Excel el listado de clientes filtrado por 'q' (nombre o cédula).
    """
    q = request.GET.get("q", "").strip()

    qs = Cliente.objects.all().order_by("nombre")
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(cedula__icontains=q))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Cabecera sin email
    ws.append(["ID", "Nombre", "Cédula", "Teléfono", "Dirección", "Creado Por"])

    for c in qs:
        ws.append(
            [
                c.id,
                c.nombre,
                c.cedula,
                c.telefono,
                c.direccion,
                c.creado_por.username if c.creado_por else "",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="clientes.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin, login_url="login")
def exportar_productos_excel(request):
    # puedes repetir el filtro por ?buscar=… si lo necesitas
    query = request.GET.get("buscar", "").strip()
    qs = (
        Product.objects.filter(
            Q(reference__icontains=query) | Q(description__icontains=query)
        )
        if query
        else Product.objects.all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Cabecera
    ws.append(["Referencia", "Descripción", "Valor Compra", "Stock"])

    for p in qs:
        ws.append(
            [
                p.reference,
                p.description,
                float(p.purchase_price),
                p.stock,
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response
